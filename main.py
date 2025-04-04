import os
import configparser
import requests
import openpyxl
import json  # Needed for Google API
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# --- Configuration ---
CONFIG_FILE = ".config"  # Path to the configuration file storing API keys
EXCEL_FILE = "prompts.xlsx"  # Path to the Excel file for prompts and results
SYSTEM_PROMPT_FILE = (
    "systemprompt.txt"  # Path to the file containing the system prompt for the AI
)
FONT_PATH = "fonts/SourceHanSans-VF.ttf.ttc"  # Path to the desired font file (used for reference)
TARGET_FONT_NAME = "Source Han Sans CN"  # Font name as recognized by Excel
API_TIMEOUT = 180  # Increased timeout for potentially multiple long calls

# --- Helper Functions ---


def read_config():
    """
    Reads all API configurations (KEY, ENDPOINT, MODEL, NAME, TYPE, ENABLED)
    from sections starting with [API_] in the .config file.
    Returns a list of dictionaries, each representing an LLM config,
    or an empty list if no valid sections are found or an error occurs.
    """
    config = configparser.ConfigParser()
    llm_configs = []
    if not os.path.exists(CONFIG_FILE):
        print(f"Error: Configuration file '{CONFIG_FILE}' not found.")
        return llm_configs
    try:
        config.read(CONFIG_FILE, encoding="utf-8")  # Specify encoding
        for section in config.sections():
            if section.startswith("API_"):
                api_config = {
                    "KEY": config.get(section, "KEY", fallback=None),
                    "ENDPOINT": config.get(section, "ENDPOINT", fallback=None),
                    "MODEL": config.get(section, "MODEL", fallback=None),
                    "NAME": config.get(
                        section, "NAME", fallback=section[4:]
                    ),  # Default name from section
                    "TYPE": config.get(
                        section, "TYPE", fallback="openai"
                    ).lower(),  # Default type is openai, ensure lowercase
                    "ENABLED": config.getboolean(section, "ENABLED", fallback=True),
                }
                # Basic validation
                if (
                    api_config["KEY"]
                    and api_config["ENDPOINT"]
                    and api_config["MODEL"]
                    and api_config["ENABLED"]
                ):
                    llm_configs.append(api_config)
                    print(
                        f"Loaded config for: {api_config['NAME']} (Type: {api_config['TYPE']}, Enabled: {api_config['ENABLED']})"
                    )
                else:
                    print(
                        f"Warning: Incomplete or disabled configuration in section '{section}'. Skipping."
                    )
        if not llm_configs:
            print(
                f"Error: No valid [API_*] sections found or configured in '{CONFIG_FILE}'."
            )
        return llm_configs
    except configparser.Error as e:
        print(f"Error reading config file '{CONFIG_FILE}': {e}")
        return []
    except Exception as e:
        print(f"An unexpected error occurred reading config: {e}")
        return []


def initialize_system_prompt():
    """
    Checks if the system prompt file exists. If not, creates an empty file.
    This ensures the script doesn't crash if the file is missing.
    """
    if not os.path.exists(SYSTEM_PROMPT_FILE):
        print(f"Creating '{SYSTEM_PROMPT_FILE}'...")
        with open(SYSTEM_PROMPT_FILE, "w", encoding="utf-8") as f:
            pass  # Create empty file
        print(f"'{SYSTEM_PROMPT_FILE}' created.")


def read_system_prompt():
    """
    Reads the content of the system prompt file.
    Ensures the file exists by calling initialize_system_prompt first.
    Returns the content as a string, or an empty string if reading fails.
    """
    initialize_system_prompt()  # Ensure file exists
    try:
        with open(SYSTEM_PROMPT_FILE, "r", encoding="utf-8") as f:
            return f.read().strip()
    except Exception as e:
        print(f"Error reading '{SYSTEM_PROMPT_FILE}': {e}")
        return ""


def apply_formatting(sheet, num_llm_cols):
    """
    Applies predefined formatting to the Excel worksheet.
    Handles dynamic number of LLM columns.
    """
    print("Applying formatting...")
    header_font = Font(name=TARGET_FONT_NAME, bold=True, size=14)
    prompt_font = Font(name=TARGET_FONT_NAME, bold=False, size=12)
    default_font = Font(name=TARGET_FONT_NAME, size=12)
    alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    llm_alignment = Alignment(
        wrap_text=True, vertical="top", horizontal="left"
    )  # Left align LLM output

    # Apply default width and header formatting
    for col_idx in range(1, sheet.max_column + 1):
        col_letter = get_column_letter(col_idx)
        # Default width, adjust specific columns below
        sheet.column_dimensions[col_letter].width = 20

        # Apply header font and alignment
        header_cell = sheet.cell(row=1, column=col_idx)
        if header_cell.value:
            header_cell.font = header_font
            header_cell.alignment = alignment

    # Specific formatting
    sheet.column_dimensions["A"].width = 40  # Wider prompt column
    sheet.column_dimensions["B"].width = 15  # Narrower flag column (was 25)

    # Set width for LLM columns (starting from column 3)
    for col_idx in range(3, 3 + num_llm_cols):
        col_letter = get_column_letter(col_idx)
        sheet.column_dimensions[col_letter].width = 60  # Wider generated text column

    for row_idx in range(1, sheet.max_row + 1):
        # Apply default font and alignment to all cells first
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value:  # Apply only if cell has content
                cell.font = default_font
                # Apply specific alignment for LLM columns
                if col_idx >= 3:
                    cell.alignment = llm_alignment
                else:
                    cell.alignment = alignment  # Center align first two columns
                # Apply specific bold font to first column (prompts), skip header
                if col_idx == 1 and row_idx > 1:
                    cell.font = prompt_font

        # Adjust row height - minimum height
        sheet.row_dimensions[row_idx].height = 21

    # Re-apply header font as it might get overwritten by the loop above
    for col_idx in range(1, sheet.max_column + 1):
        header_cell = sheet.cell(row=1, column=col_idx)
        if header_cell.value:
            header_cell.font = header_font

    print("Formatting applied.")


def initialize_excel(llm_configs):
    """
    Initializes the Excel workbook and worksheet based on LLM configs.
    - If the file doesn't exist, creates it with dynamic headers.
    - If the file exists, loads it and checks headers match config.
    Returns the workbook, sheet, and a dictionary mapping LLM names to column indices.
    Returns (None, None, None) if an error occurs.
    """
    llm_names = [cfg["NAME"] for cfg in llm_configs if cfg["ENABLED"]]
    expected_headers = ["用户提示词", "是否生成 (0 是 1 否)"] + [
        cfg["NAME"] for cfg in llm_configs if cfg["ENABLED"]
    ]
    num_llm_cols = len(llm_names)

    if not os.path.exists(EXCEL_FILE):
        print(f"Creating '{EXCEL_FILE}'...")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "AI Prompts Comparison"  # New title
        sheet.append(expected_headers)
        apply_formatting(sheet, num_llm_cols)  # Apply initial formatting
        try:
            workbook.save(EXCEL_FILE)
            print(f"'{EXCEL_FILE}' created with headers: {', '.join(expected_headers)}")
        except Exception as e:
            print(f"Error saving new Excel file '{EXCEL_FILE}': {e}")
            return None, None, None
    else:
        print(f"Loading '{EXCEL_FILE}'...")
        try:
            workbook = openpyxl.load_workbook(EXCEL_FILE)
            sheet = workbook.active
            # Check headers
            current_headers = [cell.value for cell in sheet[1]]
            if current_headers != expected_headers:
                print("Error: Header mismatch detected!")
                print(f"  Expected: {expected_headers}")
                print(f"  Found:    {current_headers}")
                print(
                    "Please ensure the Excel file headers match the configured LLMs in .config or delete the existing Excel file."
                )
                return None, None, None
            print("Headers match configuration.")
        except Exception as e:
            print(f"Error loading '{EXCEL_FILE}': {e}. Please check the file.")
            return None, None, None

    # Create mapping from LLM name to column index (1-based)
    llm_col_map = {
        cfg["NAME"]: idx + 3 for idx, cfg in enumerate(llm_configs) if cfg["ENABLED"]
    }
    return workbook, sheet, llm_col_map


def call_api(api_config, system_prompt, user_prompt):
    """
    Calls the appropriate AI API based on api_config['TYPE'].
    Handles request setup, execution, and error handling for different API types.
    Returns the generated text content or an error message string.
    """
    api_name = api_config["NAME"]
    api_type = api_config["TYPE"]
    print(
        f"  Calling API '{api_name}' (Type: {api_type}) for prompt: '{user_prompt[:30]}...'"
    )

    try:
        if api_type == "openai":
            # --- OpenAI Compatible API Call ---
            headers = {
                "Authorization": f"Bearer {api_config['KEY']}",
                "Content-Type": "application/json",
            }
            data = {
                "model": api_config["MODEL"],
                "messages": [],
            }
            # Add system prompt if provided and not empty
            if system_prompt:
                data["messages"].append({"role": "system", "content": system_prompt})
            data["messages"].append({"role": "user", "content": user_prompt})

            response = requests.post(
                f"{api_config['ENDPOINT']}/chat/completions",  # Assume /chat/completions endpoint
                headers=headers,
                json=data,
                timeout=API_TIMEOUT,
            )
            response.raise_for_status()
            result = response.json()

            # Extract content - handle potential variations in response structure
            if result.get("choices") and len(result["choices"]) > 0:
                message = result["choices"][0].get("message", {})
                content = message.get("content")
                if content:
                    print(f"  API call '{api_name}' successful.")
                    return content.strip()
                else:
                    # Handle cases like function calls if needed in the future
                    print(
                        f"  API Error ({api_name}): No 'content' found in message: {message}"
                    )
                    return f"Error ({api_name}): No content in response message."
            else:
                print(f"  API Error ({api_name}): Unexpected response format: {result}")
                return f"Error ({api_name}): Unexpected API response format."

        elif api_type == "google":
            # --- Google Gemini API Call ---
            headers = {"Content-Type": "application/json"}
            # Construct the specific URL for Google Gemini
            url = f"{api_config['ENDPOINT']}/{api_config['MODEL']}:generateContent?key={api_config['KEY']}"
            # Construct the specific JSON body for Google Gemini
            # Note: System prompt handling might differ for Gemini.
            # This basic implementation only sends the user prompt.
            # More complex scenarios might require adjusting the 'contents' structure.
            data = {"contents": [{"parts": [{"text": user_prompt}]}]}
            if system_prompt:
                # Basic system prompt integration (may need refinement based on Gemini best practices)
                data["systemInstruction"] = {"parts": [{"text": system_prompt}]}

            response = requests.post(
                url, headers=headers, json=data, timeout=API_TIMEOUT
            )
            response.raise_for_status()
            result = response.json()

            # Extract content from Google Gemini response
            if result.get("candidates") and len(result["candidates"]) > 0:
                candidate = result["candidates"][0]
                if (
                    candidate.get("content")
                    and candidate["content"].get("parts")
                    and len(candidate["content"]["parts"]) > 0
                ):
                    content = candidate["content"]["parts"][0].get("text")
                    if content:
                        print(f"  API call '{api_name}' successful.")
                        return content.strip()
                    else:
                        print(
                            f"  API Error ({api_name}): No 'text' found in content part: {candidate['content']['parts'][0]}"
                        )
                        return f"Error ({api_name}): No text in response part."
                else:
                    # Handle safety ratings, finish reasons etc. if needed
                    finish_reason = candidate.get("finishReason", "UNKNOWN")
                    safety_ratings = candidate.get("safetyRatings", [])
                    print(
                        f"  API Warning/Error ({api_name}): No content/parts found. Finish Reason: {finish_reason}. Safety: {safety_ratings}"
                    )
                    return f"Error ({api_name}): No content/parts in response. Finish: {finish_reason}"
            else:
                print(f"  API Error ({api_name}): Unexpected response format: {result}")
                return f"Error ({api_name}): No candidates in API response."

        else:
            print(
                f"  API Error ({api_name}): Unsupported API TYPE '{api_type}' in config."
            )
            return f"Error: Unsupported API type '{api_type}'"

    except requests.exceptions.Timeout:
        print(f"  API Request Error ({api_name}): Timeout after {API_TIMEOUT} seconds.")
        return f"Error ({api_name}): API request timed out."
    except requests.exceptions.RequestException as e:
        error_message = f"Error ({api_name}): API request failed."
        if e.response is not None:
            try:
                # Try to get more specific error from response body
                error_detail = e.response.json()
                error_message += (
                    f" Status: {e.response.status_code}. Detail: {error_detail}"
                )
            except json.JSONDecodeError:
                error_message += (
                    f" Status: {e.response.status_code}. Response: {e.response.text}"
                )
        else:
            error_message += f" Exception: {e}"
        print(f"  API Request Error ({api_name}): {e}")
        return error_message
    except Exception as e:
        print(f"  Error during API call processing ({api_name}): {e}")
        # Consider logging the full traceback here for debugging
        # import traceback
        # traceback.print_exc()
        return f"Error ({api_name}): Processing API response failed. {e}"


def find_column_index(sheet, header_name):
    """Finds the 1-based column index for a given header name."""
    header_row = sheet[1]
    for idx, cell in enumerate(header_row):
        if cell.value == header_name:
            return idx + 1
    return None


def read_and_process_excel(file_path):
    """
    Reads the Excel file and ensures the '是否生成 (0 是 1 否)' column has no empty values.
    If empty, it defaults to 0. Finds the column dynamically.
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        flag_header = "是否生成 (0 是 1 否)"
        col_idx = find_column_index(sheet, flag_header)

        if col_idx:
            rows_defaulted = 0
            for row in range(2, sheet.max_row + 1):  # Start from the second row
                cell = sheet.cell(row=row, column=col_idx)
                if cell.value is None:  # If the cell is empty
                    cell.value = 0  # Default to 0
                    rows_defaulted += 1
            if rows_defaulted > 0:
                print(
                    f"Defaulted {rows_defaulted} empty cells in '{flag_header}' column to 0."
                )
            wb.save(file_path)
            print(f"Checked defaults and saved: {file_path}")
        else:
            print(f"Error: Column '{flag_header}' not found in '{file_path}'!")

    except Exception as e:
        print(f"Error processing Excel file for defaults '{file_path}': {e}")


# This function seems redundant now with the check in read_and_process_excel,
# but keeping it in case it was intended for a different purpose.
# If it's truly redundant, it can be removed.
def write_excel_with_defaults(file_path):
    """
    Ensures the column '是否生成 (0 是 1 否)' has no empty values before saving.
    (Currently duplicates functionality in read_and_process_excel)
    """
    print("Warning: Function 'write_excel_with_defaults' may be redundant.")
    read_and_process_excel(file_path)  # Just call the other function for now


# --- Main Execution ---


def main():
    """
    Main function to orchestrate the script's execution.
    1. Reads all LLM configurations.
    2. Reads system prompt.
    3. Initializes or loads the Excel file, matching headers to configs.
    4. Iterates through rows, calling all configured APIs for rows marked '0'.
    5. Saves results to respective columns and updates the flag.
    6. Applies formatting and saves periodically and at the end.
    """
    print("Starting AI CellFill Excel script (Multi-LLM Version)...")

    # 1. Read Configs
    llm_configs = read_config()
    if not llm_configs:
        print("Exiting due to configuration errors.")
        return  # Stop if config is invalid or empty

    # 2. Read System Prompt
    system_prompt = read_system_prompt()
    if system_prompt:
        print(f"Using System Prompt (first 50 chars): '{system_prompt[:50]}...'")
    else:
        print("No system prompt found or read.")

    # 3. Initialize/Load Excel & Get Column Mapping
    workbook, sheet, llm_col_map = initialize_excel(llm_configs)
    if not workbook or not sheet or not llm_col_map:
        print("Exiting due to Excel initialization errors.")
        return  # Stop if Excel handling fails

    # Find the flag column index dynamically
    flag_header = "是否生成 (0 是 1 否)"
    flag_col_idx = find_column_index(sheet, flag_header)
    if not flag_col_idx:
        print(f"Critical Error: Cannot find flag column '{flag_header}'. Exiting.")
        return

    # Ensure '是否生成 (0 是 1 否)' column has defaults (run once before processing)
    read_and_process_excel(EXCEL_FILE)

    # 4. Process Rows in Excel File
    rows_processed_count = 0
    print("\nProcessing Excel rows...")
    for row_idx in range(2, sheet.max_row + 1):  # Start from row 2 (skip header)
        row_needs_processing = False
        try:
            # Get cells for the current row
            should_generate_cell = sheet.cell(row=row_idx, column=flag_col_idx)
            user_prompt_cell = sheet.cell(
                row=row_idx, column=1
            )  # Prompt is always column 1

            # Check if generation is requested (Flag Column value is 0)
            should_generate_val = should_generate_cell.value
            # Handle potential non-numeric values gracefully
            try:
                if int(should_generate_val) == 0:
                    row_needs_processing = True
                elif int(should_generate_val) == 1:
                    # Skip if Column B is explicitly set to 1
                    # print(f"Skipping row {row_idx}: Marked as '1' (Do not generate).") # Reduce verbosity
                    pass
                else:
                    print(
                        f"Skipping row {row_idx}: Invalid value '{should_generate_val}' in '{flag_header}' column (must be 0 or 1)."
                    )
            except (ValueError, TypeError):
                print(
                    f"Skipping row {row_idx}: Non-numeric value '{should_generate_val}' in '{flag_header}' column. Setting to 0."
                )
                should_generate_cell.value = 0  # Default to 0 if invalid
                row_needs_processing = True

            if row_needs_processing:
                user_prompt = (
                    str(user_prompt_cell.value).strip()
                    if user_prompt_cell.value
                    else None
                )

                if user_prompt:
                    print(
                        f"\nProcessing row {row_idx} for prompt: '{user_prompt[:50]}...'"
                    )
                    all_apis_called_for_row = True  # Assume success until an API fails

                    # Iterate through each configured LLM
                    for api_config in llm_configs:
                        if not api_config["ENABLED"]:
                            continue  # Skip disabled LLMs
                        llm_name = api_config["NAME"]
                        output_col_idx = llm_col_map.get(llm_name)

                        if not output_col_idx:
                            print(
                                f"  Error: Could not find column index for LLM '{llm_name}'. Skipping API call."
                            )
                            all_apis_called_for_row = (
                                False  # Mark row as incomplete if column missing
                            )
                            continue

                        output_cell = sheet.cell(row=row_idx, column=output_col_idx)

                        # Optional: Check if this specific cell already has content?
                        # if output_cell.value:
                        #     print(f"  Skipping API call for '{llm_name}' - cell already has content.")
                        #     continue

                        api_result = call_api(api_config, system_prompt, user_prompt)

                        # Write result (or error message) to the LLM's column
                        output_cell.value = api_result
                        if api_result and api_result.startswith("Error:"):
                            print(
                                f"  API call failed for '{llm_name}' (Row {row_idx}). Error written to cell."
                            )
                            # Decide if one failure should stop flag update? Current logic updates flag anyway.
                            # all_apis_called_for_row = False # Uncomment this if one failure should prevent flag update
                        else:
                            print(
                                f"  Result for '{llm_name}' written to row {row_idx}, column {output_col_idx}."
                            )

                    # After trying all APIs for this row, update the flag if all were attempted
                    # (Modify condition 'all_apis_called_for_row' if needed based on failure handling)
                    if all_apis_called_for_row:
                        should_generate_cell.value = 1
                        print(f"Flag updated to 1 for row {row_idx}.")
                        rows_processed_count += 1
                    else:
                        print(
                            f"Flag *not* updated for row {row_idx} due to previous errors or skipped calls."
                        )

                    # Save after each *row* is fully processed to prevent data loss
                    try:
                        # Apply formatting before saving (optional, can be slow)
                        # apply_formatting(sheet, len(llm_configs))
                        workbook.save(EXCEL_FILE)
                        print(f"'{EXCEL_FILE}' saved after processing row {row_idx}.")
                    except Exception as e:
                        print(
                            f"Error saving '{EXCEL_FILE}' after processing row {row_idx}: {e}"
                        )
                        print("Attempting to continue...")

                else:
                    # Skip if user prompt in Column A is empty
                    print(f"Skipping row {row_idx}: Empty user prompt.")
                    should_generate_cell.value = 1  # Mark as done if prompt is empty
                    print(f"Flag updated to 1 for row {row_idx} (empty prompt).")

        except (
            Exception
        ) as e:  # Catch any other unexpected errors during row processing
            print(f"Critical Error processing row {row_idx}: {e}")
            # Consider logging traceback here
            # import traceback
            # traceback.print_exc()
            # Try to mark the row with an error message in the first LLM column if possible
            try:
                first_llm_col = min(llm_col_map.values())
                sheet.cell(row=row_idx, column=first_llm_col).value = (
                    f"Error processing row: {e}"
                )
                # Also mark as 'done' to avoid retrying a broken row
                if flag_col_idx:
                    sheet.cell(row=row_idx, column=flag_col_idx).value = 1
            except:
                pass  # Ignore if we can't even write the error

    # 5. Final Save and Formatting Application
    try:
        print("\nApplying final formatting...")
        apply_formatting(
            sheet, len(llm_configs)
        )  # Ensure formatting is applied one last time
        workbook.save(EXCEL_FILE)
        print(f"Final save of '{EXCEL_FILE}' complete.")
    except Exception as e:
        print(f"Error during final save/formatting: {e}")

    print(
        f"\nScript finished. Processed {rows_processed_count} rows for all configured LLMs."
    )


if __name__ == "__main__":
    main()
