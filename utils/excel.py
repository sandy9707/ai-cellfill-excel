import os
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def apply_formatting(sheet, num_llm_cols, target_font_name="hei"):
    """
    Applies predefined formatting to the Excel worksheet.
    Handles dynamic number of LLM columns.
    """
    print("Applying formatting...")
    header_font = Font(name=target_font_name, bold=True, size=14)
    prompt_font = Font(name=target_font_name, bold=False, size=12)
    default_font = Font(name=target_font_name, size=12)
    alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    llm_alignment = Alignment(
        wrap_text=True, vertical="top", horizontal="left"
    )  # Left align LLM output

    # Apply default width and header formatting
    for col_idx in range(1, sheet.max_column + 1):
        col_letter = get_column_letter(col_idx)
        # Default width, adjust specific columns below
        sheet.column_dimensions[col_letter].width = 20

        # Apply header font and alignment
        header_cell = sheet.cell(row=1, column=col_idx)
        if header_cell.value:
            header_cell.font = header_font
            header_cell.alignment = alignment

    # Specific formatting
    sheet.column_dimensions["A"].width = 40  # Wider display column
    sheet.column_dimensions["B"].width = 40  # Wider prompt column
    sheet.column_dimensions["C"].width = 15  # Narrower flag column

    # Set width for LLM columns (starting from column 4)
    for col_idx in range(4, 4 + num_llm_cols):
        col_letter = get_column_letter(col_idx)
        sheet.column_dimensions[col_letter].width = 60  # Wider generated text column

    for row_idx in range(1, sheet.max_row + 1):
        # Apply default font and alignment to all cells first
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value:  # Apply only if cell has content
                cell.font = default_font
                # Apply specific alignment for LLM columns
                if col_idx >= 4:
                    cell.alignment = llm_alignment
                else:
                    cell.alignment = alignment  # Center align first three columns
                # Apply specific bold font to second column (prompts), skip header
                if col_idx == 2 and row_idx > 1:
                    cell.font = prompt_font

        # Adjust row height - minimum height
        sheet.row_dimensions[row_idx].height = 21

    # Re-apply header font as it might get overwritten by the loop above
    for col_idx in range(1, sheet.max_column + 1):
        header_cell = sheet.cell(row=1, column=col_idx)
        if header_cell.value:
            header_cell.font = header_font

    print("Formatting applied.")


def initialize_excel(llm_configs, excel_file="prompts.xlsx"):
    """
    Initializes the Excel workbook and worksheet based on LLM configs.
    - If the file doesn't exist, creates it with dynamic headers.
    - If the file exists, loads it and checks headers match config.
    Returns the workbook, sheet, and a dictionary mapping LLM names to column indices.
    Returns (None, None, None) if an error occurs.
    """
    llm_names = [cfg["NAME"] for cfg in llm_configs if cfg["ENABLED"]]
    expected_headers = ["用户指南", "用户提示词", "是否生成 (0 是 1 否)"] + [
        cfg["NAME"] for cfg in llm_configs if cfg["ENABLED"]
    ]
    num_llm_cols = len(llm_names)

    if not os.path.exists(excel_file):
        print(f"Creating '{excel_file}'...")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "AI Prompts Comparison"  # New title
        sheet.append(expected_headers)
        apply_formatting(sheet, num_llm_cols)  # Apply initial formatting
        try:
            workbook.save(excel_file)
            print(f"'{excel_file}' created with headers: {', '.join(expected_headers)}")
        except Exception as e:
            print(f"Error saving new Excel file '{excel_file}': {e}")
            return None, None, None
    else:
        print(f"Loading '{excel_file}'...")
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            # Check headers
            current_headers = [cell.value for cell in sheet[1]]
            if current_headers != expected_headers:
                print("Error: Header mismatch detected!")
                print(f"  Expected: {expected_headers}")
                print(f"  Found:    {current_headers}")
                print(
                    "Please ensure the Excel file headers match the configured LLMs in .config or delete the existing Excel file."
                )
                return None, None, None
            print("Headers match configuration.")
        except Exception as e:
            print(f"Error loading '{excel_file}': {e}. Please check the file.")
            return None, None, None

    # Create mapping from LLM name to column index (1-based)
    llm_col_map = {
        cfg["NAME"]: idx + 4 for idx, cfg in enumerate(llm_configs) if cfg["ENABLED"]
    }
    return workbook, sheet, llm_col_map


def find_column_index(sheet, header_name):
    """Finds the 1-based column index for a given header name."""
    header_row = sheet[1]
    for idx, cell in enumerate(header_row):
        if cell.value == header_name:
            return idx + 1
    return None


def read_and_process_excel(file_path):
    """
    Reads the Excel file and ensures the '是否生成 (0 是 1 否)' column has no empty values.
    If empty, it defaults to 0. Finds the column dynamically.
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        flag_header = "是否生成 (0 是 1 否)"
        col_idx = find_column_index(sheet, flag_header)

        if col_idx:
            rows_defaulted = 0
            for row in range(2, sheet.max_row + 1):  # Start from the second row
                cell = sheet.cell(row=row, column=col_idx)
                if cell.value is None:  # If the cell is empty
                    cell.value = 0  # Default to 0
                    rows_defaulted += 1
            if rows_defaulted > 0:
                print(
                    f"Defaulted {rows_defaulted} empty cells in '{flag_header}' column to 0."
                )
            wb.save(file_path)
            print(f"Checked defaults and saved: {file_path}")
        else:
            print(f"Error: Column '{flag_header}' not found in '{file_path}'!")

    except Exception as e:
        print(f"Error processing Excel file for defaults '{file_path}': {e}")


def write_excel_with_defaults(file_path):
    """
    Ensures the column '是否生成 (0 是 1 否)' has no empty values before saving.
    (Currently duplicates functionality in read_and_process_excel)
    """
    print("Warning: Function 'write_excel_with_defaults' may be redundant.")
    read_and_process_excel(file_path)  # Just call the other function for now